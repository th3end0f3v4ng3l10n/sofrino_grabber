#!/bin/env python3
__author__ = "Th3end0f3v4ng3l10n."
__email__  = "th3end0f3v4ng3l10n@gmail.com"
__version__ = "1.0.0"
import time

from openpyxl import load_workbook
import datetime
from openpyxl.styles import Font
import requests
from bs4 import BeautifulSoup

def hunter():
    global ids
    ids = 0


#СДелать таблицу
def table():
    date = datetime.datetime.now()
    info_for_title = '                ПРАЙС - ЛИСТ на '
    date_for_title = date.strftime('%d.%m.%y')
    year_for_title = 'г.'
    all_info = info_for_title + date_for_title + year_for_title
    print(all_info)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    sheet.merge_cells('A2:H2')
    sheet.cell(row=2,column=1).value = str(all_info)
    sheet.cell(row=2,column=1).font = Font(bold = True)
    sheet.merge_cells('A3:H3')
    sheet.cell(row =3,column=1).value = '                  ООО "ХПП "Софрино" РПЦ"'

    sheet.cell(row=4, column=1).value = '№ п/п'
    sheet.cell(row=4, column=1).font = Font(bold = True)
    sheet.cell(row=4, column=2).value = 'Артикул'
    sheet.cell(row=4, column=2).font = Font(bold = True)
    sheet.merge_cells('C4:F4')
    top = sheet['C4']
    top.value = '    Н А И М Е Н О В А Н И Е     '
    sheet.cell(row=4, column=3).font = Font(bold = True)
    ed = sheet['G4']
    ed.value = str('Ед. изм.')
    ed.font = Font(bold = True)
    fuk = sheet['H4']
    fuk.value = 'Цена розн'
    fuk.font = Font(bold = True)
    sheet.merge_cells('A5:H5')
    sheet.merge_cells('A6:H6')
    top = sheet['A6']
    top.value = '                 ИКОНОСТАСЫ  '
    sheet.merge_cells('A7:H7')
    sheet.merge_cells('A8:H8')
    top = sheet['A8']
    top.value = '                                Карниз для завесы на царские врата'
    g = 0





    wb.save('output.xlsx')

#Спарсить Карниз для завесы на царские врата
def parse_icons(url):
    global iconostasi, g, ids
    ids = 0
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })
    g = 8
    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    g+=2
    wb.save('output.xlsx')

#Спарсить Царские врата
def check_0():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Царские врата'

    wb.save('output.xlsx')
def parse_icons2(url):
    global iconostasi, g, ids
    ids = 0
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    g+=2
    wb.save('output.xlsx')

#TITLE ИКОНЫ
def check_1():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = '           ИКОНЫ'

    g+=1
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    wb.save('output.xlsx')

#Ручная шелкография
def check_2():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Ручная шелкография'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_3(url):
    global iconostasi, g, ids
    ids = 0
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#Иконы в ризе
def check_3():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Икона в ризе'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_4(url):
    global iconostasi, g, ids
    ids = 0
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#Иконы в багетной рамке
def check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Икона в багетной рамке'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_5(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')


#Икона в блистере
def check_5():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Икона в блистере'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_6(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#Иконы в деревяной рамке
def check_6():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Икона в деревяной рамке'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_7(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#Иконы в киоте
def check_7():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Икона в киоте'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_8(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#Иконы в металлической рамке
def check_8():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Икона в металлической рамке'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_9(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#Иконы в пластмассовой рамке
def check_9():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Икона в пластмассовой рамке'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_10(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#Иконы в филигранной рамке
def check_10():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Иконы в филигранной рамке'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_11(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#Иконы ламинированные
def check_11():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Иконы ламинированные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_12(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#Иконы многоместные
def check_12():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Иконы многоместные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_13(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#Иконы на гипсе
'''
def check_13():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Иконы на гипсе'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_14(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
'''
#Иконы на дереве
def check_13():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Иконы на дереве'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_15(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#Иконы на деревянном планшете
def check_14():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Иконы на деревянном планшете'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_16(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#Иконы на заказ (Именные,мерные,семейные)
def check_15():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Иконы на заказ (Именные,мерные,семейные)'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_17(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#Иконы на оргалите
def check_16():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Иконы на оргалите'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_18(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
'''
#Иконы настольные
def check_17():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Иконы настольные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_19(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#Подставки под ИКОНЫ
def check_18():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подставки под ИКОНЫ'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_20(url):
        global iconostasi, g, ids
        print(url)
        wb = load_workbook('output.xlsx')
        sheet = wb['Sheet1']
        HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
        response = requests.get(url, headers=HEADERS)
        soup = BeautifulSoup(response.content, 'html.parser')
        items = soup.findAll('div',class_='col-4 product-box')
        iconostasi = []

        for item in items:
            iconostasi.append({
                'title': item.find('a', class_='product__title').get_text(strip = True),
                'articul': item.find('div', class_='product__footer').get_text(strip = True),
                'price': item.find('div', class_='product__price__current').get_text(strip = True)
            })

        for comp in iconostasi:
            ff = []
            ff1 = []
            g += 1
            ids += 1
            ff.append(comp['articul'].split('-'))
            for i in ff[0]:
                ff1.append(i.split(' '))
            print('Articul ', ff1[0][1])
            sheet.cell(row=g, column=1).value = str(ids)
            sheet.cell(row=g,column=2).value= ff1[0][1]
            formatted = 'C{}:F{}'.format(g,g)
            sheet.merge_cells(formatted)
            formatted_2 = 'C{}'.format(g)
            top = sheet[formatted_2]
            top.value = comp['title']
            top = sheet['G{}'.format(g)]
            top.value = 'шт'
            top = sheet['H{}'.format(g)]
            top.value = comp['price']
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        wb.save('output.xlsx')

#полки под иконы
def check_19():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'полки под иконы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_21(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
'''
#складки бумажные
def check_17():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'складки бумажные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_19(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#складки в футляре
def check_18():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'складки в футляре'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_20(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#складни деревянные
def check_19():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'складни деревянные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_21(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#складни живописные
def check_20():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'складни живописные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_22(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

#царские врата(Иконы)
def check_21():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'царские врата(Иконы)'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def parse_icons_23(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')




#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#СДелать таблицу
table()

#Спарсить Карниз для завесы на царские врата
parse_icons('https://sofrino.ru/products/ikonostasy/karniz-dlya-zavesy-na-tsarskie-vrata?group_id=330&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=329&shopFilter%5Bprice_from%5D=20&shopFilter%5Bprice_to%5D=236050&shopFilter%5Bpc%5D=n')

#Спарсить Царские врата
check_0()
parse_icons2('https://sofrino.ru/products/ikonostasy/tsarskie-vrata?group_id=330&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=330&shopFilter%5Bprice_from%5D=20&shopFilter%5Bprice_to%5D=236050&shopFilter%5Bpc%5D=n')



#TITLE Иконы
check_1()


#Ручная шелкография
check_2()
hunter()
for i in range(1,3):
    parse_icons_3('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B177%5D%5B0%5D=6656&page={}'.format(i))

#Иконы в ризе
check_3()
hunter()
for i in range(1,2):
    parse_icons_4('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B0%5D=21219&page={}'.format(i))

#В багетной рамке
check_4()
hunter()
for i in range(1,4):
    parse_icons_5('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B0%5D=21233&page={}'.format(i))

#Икона в блистере
check_5()
hunter()
for i in range(1,3):
    parse_icons_6('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B0%5D=21262&page={}'.format(i))

#Иконы в деревяной рамке
check_6()
hunter()
for i in range(1,73): #73
    parse_icons_7('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B0%5D=21238&page={}'.format(i))

#Иконы в киоте
check_7()
hunter()
for i in range(1,79): #79
    parse_icons_8('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B0%5D=21218&page={}'.format(i))

#Иконы в металлической рамке
check_8()
hunter()
for i in range(1,14): #14
    parse_icons_9('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B0%5D=21245&page={}'.format(i))

#Иконы в пластмассовой рамке
check_9()
hunter()
for i in range(1,37): #37
    parse_icons_10('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B0%5D=21223&page={}'.format(i))

#Иконы в филигранной рамке
check_10()
hunter()
parse_icons_11('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B%5D=21244')

#Иконы ламинированные
check_11()
hunter()
for i in range(1,12):
    parse_icons_12('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B0%5D=21261&page={}'.format(i))

#Иконы многоместные
check_12()
hunter()
for i in range(1,3):
    parse_icons_13('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B0%5D=21230&page={}'.format(i))
#Иконы на гипсе
#NONE

#Иконы на дереве
check_13()
hunter()
for i in range(1,28): #28
    parse_icons_15('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B0%5D=21234&page={}'.format(i))


#Иконы на деревянном планшете
check_14()
hunter()
for i in range(1,71): #71
    parse_icons_16('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B0%5D=21228&page={}'.format(i))

#Иконы на заказ (Именные,мерные,семейные)
check_15()
hunter()
parse_icons_17('https://sofrino.ru/products/ikony/ikony-na-zakaz?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=519&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n')

#Иконы на оргалите
check_16()
hunter()
for i in range(1,54): #54
    parse_icons_18('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B0%5D=21236&page={}'.format(i))


'''
#Иконы настольные
check_17()
hunter()
for i in range(1,3):
    parse_icons_19(''.format(i))


#Подставки под ИКОНЫ
check_18()
hunter()
for i in range(1,3):
    parse_icons_20(''.format(i))
#полки под иконы
check_19()
hunter()
for i in range(1,3):
    parse_icons_21(''.format(i))
'''

#складки бумажные
check_17()
hunter()
parse_icons_19('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B%5D=21263')
#складки в футляре
check_18()
hunter()
for i in range(1,6): #6
    parse_icons_20('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B0%5D=21221&page=5'.format(i))
#складни деревянные
check_19()
hunter()
for i in range(1,9): #9
    parse_icons_21('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B0%5D=21240&page={}'.format(i))
#складни живописные
check_20()
hunter()
parse_icons_22('https://sofrino.ru/products/ikony?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bprice_from%5D=2870&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n&shopFilter%5Bprops%5D%5B185%5D%5B%5D=21263&shopFilter%5Bprops%5D%5B185%5D%5B%5D=21221&shopFilter%5Bprops%5D%5B185%5D%5B%5D=21240&shopFilter%5Bprops%5D%5B177%5D%5B%5D=6680')

#царские врата(Иконы)
check_21()
hunter()
parse_icons_23('https://sofrino.ru/products/ikony/ikony-na-tsarskie-vrata?group_id=299&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=520&shopFilter%5Bprice_from%5D=5&shopFilter%5Bprice_to%5D=257000&shopFilter%5Bpc%5D=n')


#TITLE КИОТЫ
def kioti():
    global g, ids
    g+=1
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = '           КИОТЫ'

    g+=1
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]


def kioto_check():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'КИОТЫ'
    g+=2
    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def kioti_parse(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

kioti()
kioto_check()
hunter()
for i in range(1,59): #59
    kioti_parse('https://sofrino.ru/poisk2?group_id=0&tag_id=0&is_search=1&phrase=%D0%BA%D0%B8%D0%BE%D1%82&shopFilter%5Bprops%5D%5B168%5D%5B0%5D=3544&shopFilter%5Bprops%5D%5B168%5D%5B1%5D=19374&shopFilter%5Bprops%5D%5B168%5D%5B2%5D=5236&shopFilter%5Bprops%5D%5B168%5D%5B3%5D=3500&shopFilter%5Bprice_from%5D=54&shopFilter%5Bprice_to%5D=579100&shopFilter%5Bpc%5D=n&page={}'.format(i))

#TITLE Печатная продукция

def produkcia():
    global g, ids
    g+=2
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = '           ПЕЧАТНАЯ ПРОДУКЦИЯ'



#Вифлеемская звезда
def prod_check_1():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Вифлеемская звезда'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def prod_parse_1(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

produkcia()
prod_check_1()
hunter()
prod_parse_1('https://sofrino.ru/products/pechatnaya-produktsiya/vifleemskaya-zvezda?group_id=301&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=445&shopFilter%5Bprice_from%5D=1&shopFilter%5Bprice_to%5D=198&shopFilter%5Bpc%5D=n')

#Грамот и благодарственные письма
def prod_check_2():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Грамот и благодарственные письма'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def prod_parse_2(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
prod_check_2()
hunter()
prod_parse_2('https://sofrino.ru/products/pechatnaya-produktsiya/gramoty-blagodarstvennye-pisma?group_id=301&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=439&shopFilter%5Bprice_from%5D=1&shopFilter%5Bprice_to%5D=198&shopFilter%5Bpc%5D=n')

#Закладки
def prod_check_3():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Закладки'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def prod_parse_3(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
prod_check_3()
hunter()
for i in range(1,4):
    prod_parse_3('https://sofrino.ru/products/pechatnaya-produktsiya/zakladki?group_id=301&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=433&shopFilter%5Bprice_from%5D=1&shopFilter%5Bprice_to%5D=198&shopFilter%5Bpc%5D=n&page={}'.format(i))


def prod_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Икона (печатное изображение)'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def prod_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
prod_check_4()
hunter()
prod_parse_4('https://sofrino.ru/products/pechatnaya-produktsiya/ikona-pechatnoe-izobrazhenie?group_id=433&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=434&shopFilter%5Bprice_from%5D=3&shopFilter%5Bprice_to%5D=5&shopFilter%5Bpc%5D=n')


def prod_check_5():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Календари'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def prod_parse_5(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
prod_check_5()
hunter()
for i in range(1,2):
    prod_parse_5('https://sofrino.ru/products/pechatnaya-produktsiya/kalendari-tserkovnye?group_id=301&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=431&shopFilter%5Bprice_from%5D=1&shopFilter%5Bprice_to%5D=198&shopFilter%5Bpc%5D=n&page=2')


def prod_check_6():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Молитвы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def prod_parse_6(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
prod_check_6()
hunter()
for i in range(1,9):
    prod_parse_6('https://sofrino.ru/products/pechatnaya-produktsiya/molitvy?page={}'.format(i))


def prod_check_7():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Открытки'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def prod_parse_7(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
prod_check_7()
hunter()
for i in range(1,5):
    prod_parse_7('https://sofrino.ru/products/pechatnaya-produktsiya/otkrytki?group_id=432&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=435&shopFilter%5Bprice_from%5D=1&shopFilter%5Bprice_to%5D=198&shopFilter%5Bpc%5D=n&page={}'.format(i))



def prod_check_8():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Поминание'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def prod_parse_8(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
prod_check_8()
hunter()
prod_parse_8('https://sofrino.ru/products/pechatnaya-produktsiya/pominanie?group_id=435&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=444&shopFilter%5Bprice_from%5D=1&shopFilter%5Bprice_to%5D=198&shopFilter%5Bpc%5D=n')

def prod_check_9():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Разное'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def prod_parse_9(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
prod_check_9()
hunter()
prod_parse_9('https://sofrino.ru/products/raznoe')

def prod_check_10():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Свидетельства'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def prod_parse_10(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
prod_check_10()
hunter()
prod_parse_10('https://sofrino.ru/products/pechatnaya-produktsiya/svidetelstvo?group_id=301&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=437&shopFilter%5Bprice_from%5D=1&shopFilter%5Bprice_to%5D=198&shopFilter%5Bpc%5D=n')

def portrety():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = '          ПОРТРЕТЫ ПАТРИАРХОВ'
    g+=2

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = '          ПОРТРЕТЫ ПАТРИАРХОВ'
def portrety_check_1():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Портреты патриархов'
    g+=1
    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def portrety_parse_1(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

portrety()
portrety_check_1()
hunter()
for i in range(1,3):
    portrety_parse_1('https://sofrino.ru/products/raznoe/portrety?page={}'.format(i))

def pos_1():
    global g, ids
    g+=2
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = '          Пошивочная продукция'
def pos_check_1():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Апостольники'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_1(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

pos_1()
pos_check_1()
hunter()
pos_parse_1('https://sofrino.ru/poisk2?phrase=%D0%90%D0%BF%D0%BE%D1%81%D1%82%D0%BE%D0%BB%D1%8C%D0%BD%D0%B8%D0%BA%D0%B8')

def pos_check_2():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'архиерейское'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_2(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')

pos_check_2()
hunter()
for i in range(1,12):
    pos_parse_2('https://sofrino.ru/poisk2?phrase=%D0%B0%D1%80%D1%85%D0%B8%D0%B5%D1%80%D0%B5%D0%B9%D1%81%D0%BA%D0%BE%D0%B5&page={}'.format(i))

def pos_check_3():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Брюки монашеские'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_3(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_3()
hunter()
for i in range(1,4):
    pos_parse_3('https://sofrino.ru/poisk2?phrase=%D0%92%D0%9E%D0%B7%D0%B4%D1%83%D1%85%D0%B0&page={}'.format(i))

def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Воскресенье христово'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%B2%D0%BE%D1%81%D0%BA%D1%80%D0%B5%D1%81%D0%B5%D0%BD%D1%8C%D0%B5+%D1%85%D1%80%D0%B8%D1%81%D1%82%D0%BE%D0%B2%D0%BE&page={}'.format(i))

def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Дароносицы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%94%D0%B0%D1%80%D0%BE%D0%BD%D0%BE%D1%81%D0%B8%D1%86%D1%8B')


def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'День ангела'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%B4%D0%B5%D0%BD%D1%8C+%D0%B0%D0%BD%D0%B3%D0%B5%D0%BB%D0%B0')

def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Диаконские'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%B4%D0%B8%D0%B0%D0%BA%D0%BE%D0%BD%D1%81%D0%BA%D0%B8%D0%B5&page={}'.format(i))

def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'для венчания'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%B4%D0%BB%D1%8F%20%D0%B2%D0%B5%D0%BD%D1%87%D0%B0%D0%BD%D0%B8%D1%8F')

def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'для крещения'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%B4%D0%BB%D1%8F+%D0%BA%D1%80%D0%B5%D1%89%D0%B5%D0%BD%D0%B8%D1%8F&page={}'.format(i))

def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'для крещения'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tags/dlya-pogrebeniya/poshivochnaya-produktsiya')




def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'для погребения'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tags/dlya-pogrebeniya/poshivochnaya-produktsiya')

def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Завесы на царские врата'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tags/dlya-pogrebeniya/poshivochnaya-produktsiya')

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Закладки'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/poshivochnaya-produktsiya?phrase=%D0%97%D0%B0%D0%BA%D0%BB%D0%B0%D0%B4%D0%BA%D0%B8')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Иерейские'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2?group_id=0&tag_id=0&is_search=1&phrase=%D0%9E%D0%91%D0%BB%D0%B0%D1%87%D0%B5%D0%BD%D0%B8%D1%8F&shopFilter%5Bprops%5D%5B168%5D%5B0%5D=3445&shopFilter%5Bprice_from%5D=500&shopFilter%5Bprice_to%5D=909000&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Илитон и плат'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/poshivochnaya-produktsiya/hramovye-oblacheniya-i-prinadlezhnosti?group_id=302&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=421&shopFilter%5Bprops%5D%5B168%5D%5B%5D=3807&shopFilter%5Bprops%5D%5B168%5D%5B%5D=3911&shopFilter%5Bprice_from%5D=220&shopFilter%5Bprice_to%5D=2480&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'камилавка'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9A%D0%B0%D0%BC%D0%B8%D0%BB%D0%B0%D0%B2%D0%BA%D0%B0')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Клобук'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9A%D0%BB%D0%BE%D0%B1%D1%83%D0%BA')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Комплект для освящения престола'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/poshivochnaya-produktsiya/hramovye-oblacheniya-i-prinadlezhnosti?group_id=302&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=421&shopFilter%5Bprops%5D%5B168%5D%5B%5D=1649&shopFilter%5Bprice_from%5D=220&shopFilter%5Bprice_to%5D=312000&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Куколь'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9A%D1%83%D0%BA%D0%BE%D0%BB%D1%8C')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Куколь'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9A%D1%83%D0%BA%D0%BE%D0%BB%D1%8C')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Мантии'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9C%D0%B0%D0%BD%D1%82%D0%B8%D0%B8')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Мантии'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9C%D0%B0%D0%BD%D1%82%D0%B8%D0%B8')
#######



###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Митры'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,14):
    pos_parse_4('https://sofrino.ru/poisk2/poshivochnaya-produktsiya?phrase=%D0%9C%D0%B8%D1%82%D1%80%D1%8B&page={}'.format(i))
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Облачения на аналой'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9E%D0%B1%D0%BB%D0%B0%D1%87%D0%B5%D0%BD%D0%B8%D1%8F+%D0%BD%D0%B0+%D0%B0%D0%BD%D0%B0%D0%BB%D0%BE%D0%B9')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Облачения на престол'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9E%D0%B1%D0%BB%D0%B0%D1%87%D0%B5%D0%BD%D0%B8%D1%8F+%D0%BD%D0%B0+%D0%BF%D1%80%D0%B5%D1%81%D1%82%D0%BE%D0%BB&page={}'.format(i))
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Комплект для храма'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/poshivochnaya-produktsiya?phrase=%D0%9A%D0%BE%D0%BC%D0%BF%D0%BB%D0%B5%D0%BA%D1%82+%D0%B4%D0%BB%D1%8F+%D1%85%D1%80%D0%B0%D0%BC%D0%B0')
#######



###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Орарь'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9E%D1%80%D0%B0%D1%80%D1%8C')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Орелецы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9E%D1%80%D0%BB%D0%B5%D1%86%D1%8B')
#######





###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Платок для храма'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BB%D0%B0%D1%82%D0%BE%D0%BA+%D0%B4%D0%BB%D1%8F+%D1%85%D1%80%D0%B0%D0%BC%D0%B0')
#######





###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Плащаницы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BB%D0%B0%D1%89%D0%B0%D0%BD%D0%B8%D1%86%D1%8B&page={}'.format(i))
#######




###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подризник'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BE%D0%B4%D1%80%D0%B8%D0%B7%D0%BD%D0%B8%D0%BA')
#######




###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подрясник'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BE%D0%B4%D1%80%D1%8F%D1%81%D0%BD%D0%B8%D0%BA&page={}'.format(i))
#######




###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подсакосник'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BE%D0%B4c%D0%B0%D0%BA%D0%BE%D1%81%D0%BD%D0%B8%D0%BA')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Полуряса'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BE%D0%BB%D1%83%D1%80%D1%8F%D1%81%D0%B0')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Поручи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BE%D1%80%D1%83%D1%87%D0%B8')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Пояс монашеский'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=Пояс+монашеский')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Протодиаконские'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D1%80%D0%BE%D1%82%D0%BE%D0%B4%D0%B8%D0%B0%D0%BA%D0%BE%D0%BD%D1%81%D0%BA%D0%B8%D0%B5')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Рождество христово'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/products/tags/rozhdestvo-hristovo/poshivochnaya-produktsiya?page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Рубашка монашеская'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A0%D1%83%D0%B1%D0%B0%D1%88%D0%BA%D0%B0%20%D0%BC%D0%BE%D0%BD%D0%B0%D1%88%D0%B5%D1%81%D0%BA%D0%B0%D1%8F')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Рушники для омовения'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A0%D1%83%D1%88%D0%BD%D0%B8%D0%BA%D0%B8+%D0%B4%D0%BB%D1%8F+%D0%BE%D0%BC%D0%BE%D0%B2%D0%B5%D0%BD%D0%B8%D1%8F')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Ряса'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A0%D1%8F%D1%81%D0%B0&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Скуфьи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A1%D0%BA%D1%83%D1%84%D1%8C%D0%B8')
#######



#TITLE ИКОНЫ
def check_1():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = '           СВЕЧИ'

    g+=1
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    wb.save('output.xlsx')








###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Стихари'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A1%D1%82%D0%B8%D1%85%D0%B0%D1%80%D0%B8')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Схима'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A1%D1%85%D0%B8%D0%BC%D0%B0')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Футляры, чехлы и вешала'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/products/poshivochnaya-produktsiya/futlyary-chehly-veshala?group_id=302&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=422&shopFilter%5Bprice_from%5D=19&shopFilter%5Bprice_to%5D=312000&shopFilter%5Bpc%5D=n'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Хитон'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A5%D0%B8%D1%82%D0%BE%D0%BD')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Хоругви'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2/poshivochnaya-produktsiya?phrase=%D0%A5%D0%BE%D1%80%D1%83%D0%B3%D0%B2%D0%B8&page={}'.format(i))
#######






#TITLE ИКОНЫ
def check_1():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = '           СВЕЧИ'

    g+=1
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    wb.save('output.xlsx')




















###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Архиерейские свечи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%90%D1%80%D1%85%D0%B8%D0%B5%D1%80%D0%B5%D0%B9%D1%81%D0%BA%D0%B8%D0%B5%20%D1%81%D0%B2%D0%B5%D1%87%D0%B8')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Венчальные свечи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%92%D0%B5%D0%BD%D1%87%D0%B0%D0%BB%D1%8C%D0%BD%D1%8B%D0%B5+%D1%81%D0%B2%D0%B5%D1%87%D0%B8&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Восковые свечи 100%'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%92%D0%BE%D1%81%D0%BA%D0%BE%D0%B2%D1%8B%D0%B5+%D1%81%D0%B2%D0%B5%D1%87%D0%B8+100%25&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Восковые свечи 50%'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%92%D0%BE%D1%81%D0%BA%D0%BE%D0%B2%D1%8B%D0%B5+%D1%81%D0%B2%D0%B5%D1%87%D0%B8+50%25')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Выносные свечи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%92%D1%8B%D0%BD%D0%BE%D1%81%D0%BD%D1%8B%D0%B5+%D1%81%D0%B2%D0%B5%D1%87%D0%B8&page={}'.format(i))
#######
#!!!!!!
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Декоративные свечи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%94%D0%B5%D0%BA%D0%BE%D1%80%D0%B0%D1%82%D0%B8%D0%B2%D0%BD%D1%8B%D0%B5+%D1%81%D0%B2%D0%B5%D1%87%D0%B8')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Диаконские свечи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%94%D0%B8%D0%B0%D0%BA%D0%BE%D0%BD%D1%81%D0%BA%D0%B8%D0%B5+%D1%81%D0%B2%D0%B5%D1%87%D0%B8&page={}'.format(i))
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Красные свечи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9A%D1%80%D0%B0%D1%81%D0%BD%D1%8B%D0%B5+%D1%81%D0%B2%D0%B5%D1%87%D0%B8&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Красные свечи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,26):
    pos_parse_4('https://sofrino.ru/products/svechi?group_id=294&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=448&shopFilter%5Bgroups%5D%5B1%5D=451&shopFilter%5Bgroups%5D%5B2%5D=447&shopFilter%5Bgroups%5D%5B3%5D=440&shopFilter%5Bgroups%5D%5B4%5D=457&shopFilter%5Bgroups%5D%5B5%5D=446&shopFilter%5Bgroups%5D%5B6%5D=450&shopFilter%5Bprops%5D%5B155%5D%5B0%5D=3460&shopFilter%5Bprice_from%5D=30&shopFilter%5Bprice_to%5D=6300&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Пасхальные свечи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/svechi?group_id=294&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=448&shopFilter%5Bgroups%5D%5B%5D=451&shopFilter%5Bgroups%5D%5B%5D=447&shopFilter%5Bgroups%5D%5B%5D=440&shopFilter%5Bgroups%5D%5B%5D=457&shopFilter%5Bgroups%5D%5B%5D=446&shopFilter%5Bgroups%5D%5B%5D=450&shopFilter%5Bprops%5D%5B168%5D%5B%5D=21296&shopFilter%5Bprice_from%5D=30&shopFilter%5Bprice_to%5D=6300&shopFilter%5Bpc%5D=n')
#######



###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Патриаршие свечи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/svechi?group_id=294&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=448&shopFilter%5Bgroups%5D%5B%5D=451&shopFilter%5Bgroups%5D%5B%5D=447&shopFilter%5Bgroups%5D%5B%5D=440&shopFilter%5Bgroups%5D%5B%5D=457&shopFilter%5Bgroups%5D%5B%5D=446&shopFilter%5Bgroups%5D%5B%5D=450&shopFilter%5Bprops%5D%5B168%5D%5B%5D=21292&shopFilter%5Bprice_from%5D=30&shopFilter%5Bprice_to%5D=6300&shopFilter%5Bpc%5D=n')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'рождественские свечи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A1%D0%B2%D0%B5%D1%87%D0%B0+%D1%80%D0%BE%D0%B6%D0%B4%D0%B5%D1%81%D1%82%D0%B2%D0%B5%D0%BD%D1%81%D0%BA%D0%B0%D1%8F&page={}'.format(i))
#######




###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'С днём ангела свечи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/svechi?group_id=294&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=448&shopFilter%5Bgroups%5D%5B%5D=451&shopFilter%5Bgroups%5D%5B%5D=447&shopFilter%5Bgroups%5D%5B%5D=440&shopFilter%5Bgroups%5D%5B%5D=457&shopFilter%5Bgroups%5D%5B%5D=446&shopFilter%5Bgroups%5D%5B%5D=450&shopFilter%5Bprops%5D%5B186%5D%5B%5D=21311&shopFilter%5Bprice_from%5D=30&shopFilter%5Bprice_to%5D=6300&shopFilter%5Bpc%5D=n')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'С днем святой троицы свечи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/svechi?group_id=294&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=448&shopFilter%5Bgroups%5D%5B%5D=451&shopFilter%5Bgroups%5D%5B%5D=447&shopFilter%5Bgroups%5D%5B%5D=440&shopFilter%5Bgroups%5D%5B%5D=457&shopFilter%5Bgroups%5D%5B%5D=446&shopFilter%5Bgroups%5D%5B%5D=450&shopFilter%5Bprops%5D%5B186%5D%5B%5D=21307&shopFilter%5Bprice_from%5D=30&shopFilter%5Bprice_to%5D=6300&shopFilter%5Bpc%5D=n')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'С днем святой троицы свечи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/svechi?group_id=294&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=448&shopFilter%5Bgroups%5D%5B%5D=451&shopFilter%5Bgroups%5D%5B%5D=447&shopFilter%5Bgroups%5D%5B%5D=440&shopFilter%5Bgroups%5D%5B%5D=457&shopFilter%5Bgroups%5D%5B%5D=446&shopFilter%5Bgroups%5D%5B%5D=450&shopFilter%5Bprops%5D%5B186%5D%5B%5D=21307&shopFilter%5Bprice_from%5D=30&shopFilter%5Bprice_to%5D=6300&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Свечи для домашней молитвы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,9):
    pos_parse_4('https://sofrino.ru/products/svechi?group_id=294&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=457&shopFilter%5Bgroups%5D%5B1%5D=450&shopFilter%5Bprice_from%5D=40&shopFilter%5Bprice_to%5D=4230&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######
#!!!!!
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Свечи для пасхального трехсвечника'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,9):
    pos_parse_4('https://sofrino.ru/products/svechi?group_id=294&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=448&shopFilter%5Bgroups%5D%5B%5D=451&shopFilter%5Bgroups%5D%5B%5D=447&shopFilter%5Bgroups%5D%5B%5D=440&shopFilter%5Bgroups%5D%5B%5D=457&shopFilter%5Bgroups%5D%5B%5D=446&shopFilter%5Bgroups%5D%5B%5D=450&shopFilter%5Bprops%5D%5B168%5D%5B%5D=21296&shopFilter%5Bprice_from%5D=30&shopFilter%5Bprice_to%5D=6300&shopFilter%5Bpc%5D=n')
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Свечи для домашней молитвы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,22):
    pos_parse_4('https://sofrino.ru/products/svechi?group_id=294&tag_id=0&is_search=&phrase=&shopFilter[groups][0]=448&shopFilter[groups][1]=451&shopFilter[groups][2]=447&shopFilter[groups][3]=440&shopFilter[groups][4]=457&shopFilter[groups][5]=446&shopFilter[groups][6]=450&shopFilter[props][168][0]=3461&shopFilter[price_from]=30&shopFilter[price_to]=6300&shopFilter[pc]=n&page={}'.format(i))
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Цветные свечи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,9):
    pos_parse_4('https://sofrino.ru/products/svechi/tsvetnye-svechi?group_id=294&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=446&shopFilter%5Bprops%5D%5B168%5D%5B%5D=3461&shopFilter%5Bprice_from%5D=30&shopFilter%5Bprice_to%5D=4230&shopFilter%5Bpc%5D=n')
#######

#####TITLE######
#TITLE ИКОНЫ
def check_1():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = '           Cувенирная продукция'

    g+=1
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    wb.save('output.xlsx')

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Картина Церковная'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9A%D0%B0%D1%80%D1%82%D0%B8%D0%BD%D0%B0')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Скульптуры и статуэтки '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/suveniry/skulptury-i-statuetki?group_id=295&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=311&shopFilter%5Bprice_from%5D=33&shopFilter%5Bprice_to%5D=241000&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Cувениры'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/products/suveniry/suveniry?group_id=295&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=312&shopFilter%5Bprice_from%5D=33&shopFilter%5Bprice_to%5D=241000&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Чётки'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A7%D0%B5%D1%82%D0%BA%D0%B8')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Яйца пасхальные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,9):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%AF%D0%B9%D1%86%D0%B0+%D0%BF%D0%B0%D1%81%D1%85%D0%B0%D0%BB%D1%8C%D0%BD%D1%8B%D0%B5&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Ящик специальный'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%AF%D1%89%D0%B8%D0%BA+%D1%81%D0%BF%D0%B5%D1%86%D0%B8%D0%B0%D0%BB%D1%8C%D0%BD%D1%8B%D0%B9')
#######



#TITLE ИКОНЫ
def check_1():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = '           ХРАМОВАЯ МЕБЕЛЬ'

    g+=1
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    wb.save('output.xlsx')

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Кресло-троны'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/products/hramovaya-mebel/kreslo-trony?page=2'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Седалища'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/products/hramovaya-mebel/sedalischa?group_id=323&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=325&shopFilter%5Bprice_from%5D=7050&shopFilter%5Bprice_to%5D=1124550&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Стасидии '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/hramovaya-mebel/stasidii?group_id=325&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=324&shopFilter%5Bprice_from%5D=7050&shopFilter%5Bprice_to%5D=1124550&shopFilter%5Bpc%5D=n')
#######



###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Столы литийные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/hramovaya-mebel/stoly-litiynye?group_id=325&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=322&shopFilter%5Bprice_from%5D=181600&shopFilter%5Bprice_to%5D=379950&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Ящики свечные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/hramovaya-mebel/yaschiki-svechnye?group_id=322&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=314&shopFilter%5Bprice_from%5D=7050&shopFilter%5Bprice_to%5D=1124550&shopFilter%5Bpc%5D=n')
#######

#TITLE ИКОНЫ
def check_1():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = '           ЦЕРКОВНАЯ УТВАРЬ'

    g+=1
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    wb.save('output.xlsx')




###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Аналои'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%90%D0%BD%D0%B0%D0%BB%D0%BE%D0%B8')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Апостолы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%90%D0%BF%D0%BE%D1%81%D1%82%D0%BE%D0%BB%D1%8B')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Баки для святой воды'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/baki-dlya-svyatoy-vody?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=402&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Блюда всенощные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/blyuda-vsenoschnye?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=364&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Блюда для приготовления Агнца '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/blyuda-dlya-prigotovleniya-agntsa?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=412&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Венцы '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/ventsy?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=365&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Вкладыш в кадило'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%92%D0%BA%D0%BB%D0%B0%D0%B4%D1%8B%D1%88%20%D0%B2%20%D0%BA%D0%B0%D0%B4%D0%B8%D0%BB%D0%BE')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Чаши водосвятные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/chashi-vodosvyatnye?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=404&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Вывески церковные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/vyveski-tserkovnye?group_id=404&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=367&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Голгофы на жертвенник'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/golgofy?group_id=415&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=415&shopFilter%5Bprops%5D%5B168%5D%5B%5D=2567&shopFilter%5Bprice_from%5D=280&shopFilter%5Bprice_to%5D=641500&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Голгофа напольная '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/golgofy?group_id=415&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=415&shopFilter%5Bprops%5D%5B168%5D%5B%5D=35563&shopFilter%5Bprice_from%5D=280&shopFilter%5Bprice_to%5D=641500&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Голгофы '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/golgofy?group_id=415&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=415&shopFilter%5Bprice_from%5D=62800&shopFilter%5Bprice_to%5D=641500&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Гробницы, раки  '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/grobnitsy-raki?group_id=415&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=366&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Дароносицы  '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/daronositsy?group_id=415&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=379&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Дарохранительницы   '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/darohranitelnitsy?group_id=415&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=368&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Дикирии-трикирии   '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/dikirii-trikirii?group_id=368&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=396&shopFilter%5Bprice_from%5D=2330&shopFilter%5Bprice_to%5D=575250&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Дискос'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,5):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%94%D0%B8%D1%81%D0%BA%D0%BE%D1%81&page={}'.format(i))
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Древки и оглавия для хоругви и крест-икон'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/drevki-i-oglaviya-dlya-horugvi-i-krest-ikon?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=354&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Евангелия'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%95%D0%B2%D0%B0%D0%BD%D0%B3%D0%B5%D0%BB%D0%B8%D1%8F&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Евхаристический набор и принадлежности '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,8):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/evharisticheskiy-nabor-i-prinadlezhnosti?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=407&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Жезлы '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/zhezly-posohi?group_id=407&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=369&shopFilter%5Bprice_from%5D=1080&shopFilter%5Bprice_to%5D=412950&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Жертвенники '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/zhertvenniki-oblacheniya-na-zhertvennik?group_id=369&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=389&shopFilter%5Bprice_from%5D=5560&shopFilter%5Bprice_to%5D=681000&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Звездица '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%97%D0%B2%D0%B5%D0%B7%D0%B4%D0%B8%D1%86%D0%B0')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Кадило '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9A%D0%B0%D0%B4%D0%B8%D0%BB%D0%BE&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Кандило '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9A%D0%B0%D0%BD%D0%B4%D0%B8%D0%BB%D0%BE')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Кануны '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9A%D0%B0%D0%BD%D1%83%D0%BD%D1%8B')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Ковчег '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,9):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9A%D0%BE%D0%B2%D1%87%D0%B5%D0%B3&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Ковчеги для преждеосвященных даров  '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar/kovchegi-dlya-prezhdeosvyaschennyh-darov?group_id=303&tag_id=0&is_search=1&phrase=%D0%9A%D0%BE%D0%B2%D1%87%D0%B5%D0%B3&shopFilter%5Bgroups%5D%5B%5D=413&shopFilter%5Bprice_from%5D=321&shopFilter%5Bprice_to%5D=462000&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Ковши'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9A%D0%BE%D0%B2%D1%88%D0%B8')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Копие'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9A%D0%BE%D0%BF%D0%B8%D0%B5')
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Крест водосвятный'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9A%D1%80%D0%B5%D1%81%D1%82+%D0%B2%D0%BE%D0%B4%D0%BE%D1%81%D0%B2%D1%8F%D1%82%D0%BD%D1%8B%D0%B9')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Крест выносной'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?group_id=0&tag_id=0&is_search=1&phrase=%D0%9A%D1%80%D0%B5%D1%81%D1%82+%D0%B2%D1%8B%D0%BD%D0%BE%D1%81%D0%BD%D0%BE%D0%B9&shopFilter%5Bprops%5D%5B168%5D%5B%5D=2729&shopFilter%5Bprice_from%5D=1900&shopFilter%5Bprice_to%5D=77500&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Крест деревянный'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9A%D1%80%D0%B5%D1%81%D1%82+%D0%B4%D0%B5%D1%80%D0%B5%D0%B2%D1%8F%D0%BD%D0%BD%D1%8B%D0%B9&page={}'.format(i))
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Крест для домашней молитвы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9A%D1%80%D0%B5%D1%81%D1%82+%D0%B4%D0%BB%D1%8F+%D0%B4%D0%BE%D0%BC%D0%B0%D1%88%D0%BD%D0%B5%D0%B9+%D0%BC%D0%BE%D0%BB%D0%B8%D1%82%D0%B2%D1%8B')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Крест на купол'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9A%D1%80%D0%B5%D1%81%D1%82+%D0%BD%D0%B0+%D0%BA%D1%83%D0%BF%D0%BE%D0%BB')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Кресты наперсные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/kresty-na-mitru-klobuk-skufyu?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=410&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Кресты напрестольные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,10):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/kresty-naprestolnye?group_id=385&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=406&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Крестики нательные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,5):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/krestiki-natelnye?group_id=406&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=397&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Кресты погребальные, могильные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/kresty-pogrebalnye-mogilnye?group_id=397&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=490&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'крест требный'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%BA%D1%80%D0%B5%D1%81%D1%82+%D1%82%D1%80%D0%B5%D0%B1%D0%BD%D1%8B%D0%B9')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'крест требный'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%BA%D1%80%D0%B5%D1%81%D1%82+%D1%82%D1%80%D0%B5%D0%B1%D0%BD%D1%8B%D0%B9')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = ' Крест-иконы запрестольные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar/krest-ikony-zaprestolnye?group_id=303&tag_id=0&is_search=1&phrase=%D0%BA%D1%80%D0%B5%D1%81%D1%82-%D0%B8%D0%BA%D0%BE%D0%BD%D0%B0&shopFilter%5Bgroups%5D%5B0%5D=392&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=639550&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = ' Кронштейн для лампады '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2?group_id=0&tag_id=0&is_search=1&phrase=%D0%9A%D1%80%D0%BE%D0%BD%D1%88%D1%82%D0%B5%D0%B9%D0%BD+%D0%B4%D0%BB%D1%8F+%D0%BB%D0%B0%D0%BC%D0%BF%D0%B0%D0%B4%D1%8B&shopFilter%5Bprops%5D%5B168%5D%5B0%5D=2052&shopFilter%5Bprice_from%5D=440&shopFilter%5Bprice_to%5D=34200&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Кропило'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9A%D1%80%D0%BE%D0%BF%D0%B8%D0%BB%D0%BE')
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Кружка для пожертвования'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9A%D1%80%D1%83%D0%B6%D0%BA%D0%B0+%D0%B4%D0%BB%D1%8F+%D0%BF%D0%BE%D0%B6%D0%B5%D1%80%D1%82%D0%B2%D0%BE%D0%B2%D0%B0%D0%BD%D0%B8%D1%8F')
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Крышки панихидного стола'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?group_id=0&tag_id=0&is_search=1&phrase=%D0%9A%D1%80%D1%8B%D1%88%D0%BA%D0%B8+%D0%BF%D0%B0%D0%BD%D0%B8%D1%85%D0%B8%D0%B4%D0%BD%D0%BE%D0%B3%D0%BE+%D1%81%D1%82%D0%BE%D0%BB%D0%B0&shopFilter%5Bprops%5D%5B168%5D%5B%5D=3343&shopFilter%5Bprice_from%5D=17400&shopFilter%5Bprice_to%5D=148000&shopFilter%5Bpc%5D=n')
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'кувшин для омовения'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%BA%D1%83%D0%B2%D1%88%D0%B8%D0%BD+%D0%B4%D0%BB%D1%8F+%D0%BE%D0%BC%D0%BE%D0%B2%D0%B5%D0%BD%D0%B8%D1%8F')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'купели'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%BA%D1%83%D0%BF%D0%B5%D0%BB%D0%B8')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Ладан'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar/ladan?group_id=303&tag_id=0&is_search=1&phrase=%D0%9B%D0%B0%D0%B4%D0%B0%D0%BD&shopFilter%5Bgroups%5D%5B%5D=359&shopFilter%5Bprice_from%5D=25&shopFilter%5Bprice_to%5D=56400&shopFilter%5Bpc%5D=n')
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Ладаница'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar/ladanitsa?group_id=303&tag_id=0&is_search=1&phrase=%D0%9B%D0%B0%D0%B4%D0%B0%D0%BD&shopFilter%5Bgroups%5D%5B0%5D=408&shopFilter%5Bprice_from%5D=25&shopFilter%5Bprice_to%5D=56400&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Лампада неугасимая'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9B%D0%B0%D0%BC%D0%BF%D0%B0%D0%B4%D0%B0+%D0%BD%D0%B5%D1%83%D0%B3%D0%B0%D1%81%D0%B8%D0%BC%D0%B0%D1%8F')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Лампада на кронштейне'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?group_id=0&tag_id=0&is_search=1&phrase=%D0%9B%D0%B0%D0%BC%D0%BF%D0%B0%D0%B4%D0%B0+%D0%BD%D0%B0+%D0%BA%D1%80%D0%BE%D0%BD%D1%88%D1%82%D0%B5%D0%B9%D0%BD%D0%B5&shopFilter%5Bprops%5D%5B168%5D%5B%5D=2743&shopFilter%5Bprice_from%5D=440&shopFilter%5Bprice_to%5D=34200&shopFilter%5Bpc%5D=n')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Лампада настенные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9B%D0%B0%D0%BC%D0%BF%D0%B0%D0%B4%D0%B0+%D0%BD%D0%B0%D1%81%D1%82%D0%B5%D0%BD%D0%BD%D1%8B%D0%B5&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Лампада настенные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9B%D0%B0%D0%BC%D0%BF%D0%B0%D0%B4%D0%B0+%D0%BD%D0%B0%D1%81%D1%82%D0%B5%D0%BD%D0%BD%D1%8B%D0%B5&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Лампада настольная'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9B%D0%B0%D0%BC%D0%BF%D0%B0%D0%B4%D0%B0+%D0%BD%D0%B0%D1%81%D1%82%D0%BE%D0%BB%D1%8C%D0%BD%D0%B0%D1%8F&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Лампада подвесные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,11):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9B%D0%B0%D0%BC%D0%BF%D0%B0%D0%B4%D0%B0+%D0%BF%D0%BE%D0%B4%D0%B2%D0%B5%D1%81%D0%BD%D1%8B%D0%B5&page={}'.format(i))
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Лжицы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9B%D0%B6%D0%B8%D1%86%D1%8B')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Масле лампадное'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9C%D0%B0%D1%81%D0%BB%D0%B5+%D0%BB%D0%B0%D0%BC%D0%BF%D0%B0%D0%B4%D0%BD%D0%BE%D0%B5')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Масле освященное'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9C%D0%B0%D1%81%D0%BB%D0%B5+%D0%BE%D1%81%D0%B2%D1%8F%D1%89%D0%B5%D0%BD%D0%BD%D0%BE%D0%B5')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Мощевики'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,5):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9C%D0%BE%D1%89%D0%B5%D0%B2%D0%B8%D0%BA%D0%B8&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'наборы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%BD%D0%B0%D0%B1%D0%BE%D1%80&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'накладки на алтарную дверь'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%BD%D0%B0%D0%BA%D0%BB%D0%B0%D0%B4%D0%BA%D0%B8+%D0%BD%D0%B0+%D0%B0%D0%BB%D1%82%D0%B0%D1%80%D0%BD%D1%83%D1%8E+%D0%B4%D0%B2%D0%B5%D1%80%D1%8C')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'облачение на жертвенник'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?group_id=0&tag_id=0&is_search=1&phrase=%D0%BE%D0%B1%D0%BB%D0%B0%D1%87%D0%B5%D0%BD%D0%B8%D0%B5+%D0%BD%D0%B0+%D0%B6%D0%B5%D1%80%D1%82%D0%B2%D0%B5%D0%BD%D0%BD%D0%B8%D0%BA&shopFilter%5Bprops%5D%5B168%5D%5B%5D=1639&shopFilter%5Bprice_from%5D=49800&shopFilter%5Bprice_to%5D=681000&shopFilter%5Bpc%5D=n')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'облачение на жертвенник'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?group_id=0&tag_id=0&is_search=1&phrase=%D0%BE%D0%B1%D0%BB%D0%B0%D1%87%D0%B5%D0%BD%D0%B8%D0%B5+%D0%BD%D0%B0+%D0%B6%D0%B5%D1%80%D1%82%D0%B2%D0%B5%D0%BD%D0%BD%D0%B8%D0%BA&shopFilter%5Bprops%5D%5B168%5D%5B%5D=1639&shopFilter%5Bprice_from%5D=49800&shopFilter%5Bprice_to%5D=681000&shopFilter%5Bpc%5D=n')
#######


###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'облачение на престол'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar/prestoly-oblacheniya-na-prestol?group_id=303&tag_id=0&is_search=1&phrase=%D0%BE%D0%B1%D0%BB%D0%B0%D1%87%D0%B5%D0%BD%D0%B8%D0%B5+%D0%BD%D0%B0+%D0%BF%D1%80%D0%B5%D1%81%D1%82%D0%BE%D0%BB&shopFilter%5Bgroups%5D%5B0%5D=388&shopFilter%5Bprice_from%5D=19400&shopFilter%5Bprice_to%5D=909000&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Панагии'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,7):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9F%D0%B0%D0%BD%D0%B0%D0%B3%D0%B8%D0%B8&page={}'.format(i))
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Паникадила'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,5):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%B0%D0%BD%D0%B8%D0%BA%D0%B0%D0%B4%D0%B8%D0%BB%D0%B0&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Панихидные столы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2?group_id=0&tag_id=0&is_search=1&phrase=%D0%9F%D0%B0%D0%BD%D0%B8%D1%85%D0%B8%D0%B4%D0%BD%D1%8B%D0%B5+%D1%81%D1%82%D0%BE%D0%BB%D1%8B&shopFilter%5Bprops%5D%5B168%5D%5B0%5D=3323&shopFilter%5Bprice_from%5D=17400&shopFilter%5Bprice_to%5D=546000&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'пасочница'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%BF%D0%B0%D1%81%D0%BE%D1%87%D0%BD%D0%B8%D1%86%D0%B0')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'печати'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%BF%D0%B5%D1%87%D0%B0%D1%82%D1%8C&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подиум-кафедра'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BE%D0%B4%D0%B8%D1%83%D0%BC-%D0%BA%D0%B0%D1%84%D0%B5%D0%B4%D1%80%D0%B0')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Пасхальный подсвечник'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9F%D0%B0%D1%81%D1%85%D0%B0%D0%BB%D1%8C%D0%BD%D1%8B%D0%B9+%D0%BF%D0%BE%D0%B4%D1%81%D0%B2%D0%B5%D1%87%D0%BD%D0%B8%D0%BA')
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подсвечники напольные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,6):
    pos_parse_4('https://sofrino.ru/poisk2?group_id=0&tag_id=0&is_search=1&phrase=%D0%9F%D0%BE%D0%B4%D1%81%D0%B2%D0%B5%D1%87%D0%BD%D0%B8%D0%BA%D0%B8+%D0%BD%D0%B0%D0%BF%D0%BE%D0%BB%D1%8C%D0%BD%D1%8B%D0%B5&shopFilter%5Bprops%5D%5B168%5D%5B0%5D=3026&shopFilter%5Bprice_from%5D=4900&shopFilter%5Bprice_to%5D=784000&shopFilter%5Bpc%5D=n&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подсвечники настольные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,6):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BE%D0%B4%D1%81%D0%B2%D0%B5%D1%87%D0%BD%D0%B8%D0%BA%D0%B8+%D0%BD%D0%B0%D1%81%D1%82%D0%BE%D0%BB%D1%8C%D0%BD%D1%8B%D0%B5&page={}'.format(i))
#######

###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подсвечники ручные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9F%D0%BE%D0%B4%D1%81%D0%B2%D0%B5%D1%87%D0%BD%D0%B8%D0%BA%D0%B8+%D1%80%D1%83%D1%87%D0%BD%D1%8B%D0%B5')
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подставка под блюдо всенощное'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BE%D0%B4%D1%81%D1%82%D0%B0%D0%B2%D0%BA%D0%B0+%D0%BF%D0%BE%D0%B4+%D0%B1%D0%BB%D1%8E%D0%B4%D0%BE+%D0%B2%D1%81%D0%B5%D0%BD%D0%BE%D1%89%D0%BD%D0%BE%D0%B5')
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подставка под дарохранительницу'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BE%D0%B4%D1%81%D1%82%D0%B0%D0%B2%D0%BA%D0%B0+%D0%BF%D0%BE%D0%B4+%D0%B4%D0%B0%D1%80%D0%BE%D1%85%D1%80%D0%B0%D0%BD%D0%B8%D1%82%D0%B5%D0%BB%D1%8C%D0%BD%D0%B8%D1%86%D1%83&page={}'.format(i))
#######
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подставка под дикирий-трикирий'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BE%D0%B4%D1%81%D1%82%D0%B0%D0%B2%D0%BA%D0%B0+%D0%BF%D0%BE%D0%B4+%D0%B4%D0%B8%D0%BA%D0%B8%D1%80%D0%B8%D0%B9-%D1%82%D1%80%D0%B8%D0%BA%D0%B8%D1%80%D0%B8%D0%B9')
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подставка под жезл, посох'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BE%D0%B4%D1%81%D1%82%D0%B0%D0%B2%D0%BA%D0%B0+%D0%BF%D0%BE%D0%B4+%D0%B6%D0%B5%D0%B7%D0%BB%2C+%D0%BF%D0%BE%D1%81%D0%BE%D1%85')
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подставка под жезл, посох'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BE%D0%B4%D1%81%D1%82%D0%B0%D0%B2%D0%BA%D0%B0+%D0%BF%D0%BE%D0%B4+%D0%B6%D0%B5%D0%B7%D0%BB%2C+%D0%BF%D0%BE%D1%81%D0%BE%D1%85')
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подставка под кадило'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BE%D0%B4%D1%81%D1%82%D0%B0%D0%B2%D0%BA%D0%B0+%D0%BF%D0%BE%D0%B4+%D0%BA%D0%B0%D0%B4%D0%B8%D0%BB%D0%BE')
#######
###################################BLOCK########################
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подставка под ковчег для святых мощей'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BE%D0%B4%D1%81%D1%82%D0%B0%D0%B2%D0%BA%D0%B0+%D0%BF%D0%BE%D0%B4+%D0%BA%D0%BE%D0%B2%D1%87%D0%B5%D0%B3')
#######

def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подставка под крест-икону, хоругви, рипиду'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D0%BE%D0%B4%D1%81%D1%82%D0%B0%D0%B2%D0%BA%D0%B0+%D0%BF%D0%BE%D0%B4+%D0%BA%D1%80%D0%B5%D1%81%D1%82-%D0%B8%D0%BA%D0%BE%D0%BD%D1%83%2C+%D1%85%D0%BE%D1%80%D1%83%D0%B3%D0%B2%D0%B8%2C+%D1%80%D0%B8%D0%BF%D0%B8%D0%B4%D1%83')
#######
#######

def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подставка под пасхальные яйца'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9F%D0%BE%D0%B4%D1%81%D1%82%D0%B0%D0%B2%D0%BA%D0%B0+%D0%BF%D0%BE%D0%B4+%D0%BF%D0%B0%D1%81%D1%85%D0%B0%D0%BB%D1%8C%D0%BD%D1%8B%D0%B5+%D1%8F%D0%B9%D1%86%D0%B0')
#######
#######

def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Поплавок для лампад '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?group_id=0&tag_id=0&is_search=1&phrase=%D0%9F%D0%BE%D0%BF%D0%BB%D0%B0%D0%B2%D0%BE%D0%BA&shopFilter%5Bprops%5D%5B168%5D%5B%5D=1773&shopFilter%5Bprice_from%5D=24&shopFilter%5Bprice_to%5D=26&shopFilter%5Bpc%5D=n')
#######

#######

def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Посохи'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9F%D0%BE%D1%81%D0%BE%D1%85%D0%B8&page={}'.format(i))
#######
#######

def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Потир-чаши'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,7):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9F%D0%BE%D1%82%D0%B8%D1%80-%D1%87%D0%B0%D1%88%D0%B8&page={}'.format(i))
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'престолы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%9F%D1%80%D0%B5%D1%81%D1%82%D0%BE%D0%BB%D1%8B&page={}'.format(i))
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Приборы для соборования'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%9F%D1%80%D0%B8%D0%B1%D0%BE%D1%80%D1%8B+%D0%B4%D0%BB%D1%8F+%D1%81%D0%BE%D0%B1%D0%BE%D1%80%D0%BE%D0%B2%D0%B0%D0%BD%D0%B8%D1%8F')
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Разное'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%91%D0%B0%D0%BA%20%D0%BF%D0%BE%D0%B4%20%D0%BB%D0%B0%D0%BC%D0%BF%D0%B0%D0%B4%D0%BD%D0%BE%D0%B5%20%D0%BC%D0%B0%D1%81%D0%BB%D0%BE%2030%D0%BB')
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Рамки на митру'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A0%D0%B0%D0%BC%D0%BA%D0%B8+%D0%BD%D0%B0+%D0%BC%D0%B8%D1%82%D1%80%D1%83')
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Распродажа'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,6):
    pos_parse_4('https://sofrino.ru/products/tags/aktsiya?page={}'.format(i))
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Решетки на солею'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A0%D0%B5%D1%88%D0%B5%D1%82%D0%BA%D0%B8+%D0%BD%D0%B0+%D1%81%D0%BE%D0%BB%D0%B5%D1%8E&page={}'.format(i))
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Рипиды'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A0%D0%B8%D0%BF%D0%B8%D0%B4%D1%8B&page={}'.format(i))
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Свечегаситель'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A1%D0%B2%D0%B5%D1%87%D0%B5%D0%B3%D0%B0%D1%81%D0%B8%D1%82%D0%B5%D0%BB%D1%8C')
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Семисвечники'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A1%D0%B5%D0%BC%D0%B8%D1%81%D0%B2%D0%B5%D1%87%D0%BD%D0%B8%D0%BA%D0%B8&page={}'.format(i))
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Сень надпрестольная'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A1%D0%B5%D0%BD%D1%8C+%D0%BD%D0%B0%D0%B4%D0%BF%D1%80%D0%B5%D1%81%D1%82%D0%BE%D0%BB%D1%8C%D0%BD%D0%B0%D1%8F')
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Сосуды для елея'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/sosudy-dlya-eleya?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=399&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Сосуды для миро'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/sosudy-dlya-miro?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=400&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Сосуд для причастия'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/tserkovnaya-utvar?phrase=%D0%A1%D0%BE%D1%81%D1%83%D0%B4+%D0%B4%D0%BB%D1%8F+%D0%BF%D1%80%D0%B8%D1%87%D0%B0%D1%81%D1%82%D0%B8%D1%8F')
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Стаканы для лампад '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/stakany-dlya-lampad?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=504&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Стрючци  '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/stryuchtsi?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=360&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Трехсвечники   '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/trehsvechniki?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=409&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')
#######
#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Уголь кадильный'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/ugol-kadilnyy?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=362&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')

#######

#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'фитиль для лампад'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D1%84%D0%B8%D1%82%D0%B8%D0%BB%D1%8C+%D0%B4%D0%BB%D1%8F+%D0%BB%D0%B0%D0%BC%D0%BF%D0%B0%D0%B4')

#######

#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'фонари'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/fonari-pashalnye?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=386&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')

#######


#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Хоругви'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/horugvi?group_id=303&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=394&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n&page={}'.format(i))

#######

#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Цепи для крестов и панагий '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/tsepi-dlya-krestov-i-panagiy?group_id=394&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=352&shopFilter%5Bprice_from%5D=6850&shopFilter%5Bprice_to%5D=442000&shopFilter%5Bpc%5D=n')

#######
#######

#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Чаши для святых частиц '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tserkovnaya-utvar/chashi-dlya-svyatyh-chastits?group_id=394&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B%5D=398&shopFilter%5Bprice_from%5D=6&shopFilter%5Bprice_to%5D=3690000&shopFilter%5Bpc%5D=n')

#######


#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Чиновник  '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?group_id=0&tag_id=0&is_search=1&phrase=%D0%A7%D0%B8%D0%BD%D0%BE%D0%B2%D0%BD%D0%B8%D0%BA&shopFilter%5Bprops%5D%5B168%5D%5B%5D=2410&shopFilter%5Bprice_from%5D=8800&shopFilter%5Bprice_to%5D=47600&shopFilter%5Bpc%5D=n')

#######


#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Штандарты  '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A8%D1%82%D0%B0%D0%BD%D0%B4%D0%B0%D1%80%D1%82%D1%8B')

#######

#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Ящик для свечей и огарков   '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?group_id=0&tag_id=0&is_search=1&phrase=%D1%8F%D1%89%D0%B8%D0%BA+%D0%B4%D0%BB%D1%8F+%D1%81%D0%B2%D0%B5%D1%87%D0%B5%D0%B9&shopFilter%5Bprops%5D%5B168%5D%5B%5D=654&shopFilter%5Bprice_from%5D=4000&shopFilter%5Bprice_to%5D=1974000&shopFilter%5Bpc%5D=n')

#######

#######
def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'ящики крестильные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D1%8F%D1%89%D0%B8%D0%BA%D0%B8+%D0%BA%D1%80%D0%B5%D1%81%D1%82%D0%B8%D0%BB%D1%8C%D0%BD%D1%8B%D0%B5')

#######
#TITLE ИКОНЫ
def check_1():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = '           ЮВЕЛИРНАЯ ПРОДУКЦИЯ'

    g+=1
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    wb.save('output.xlsx')

def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'БЛюдо под дискос'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%91%D0%BB%D1%8E%D0%B4%D0%B0%20%D0%BF%D0%BE%D0%B4%20%D0%B4%D0%B8%D1%81%D0%BA%D0%BE%D1%81')


def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'дикирии-трикирии'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%B4%D0%B8%D0%BA%D0%B8%D1%80%D0%B8%D0%B8-%D1%82%D1%80%D0%B8%D0%BA%D0%B8%D1%80%D0%B8%D0%B8&page={}'.format(i))





def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Дискосы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/poisk2/yuvelirnaya-produktsiya?phrase=%D0%94%D0%B8%D1%81%D0%BA%D0%BE%D1%81%D1%8B&page={}'.format(i))



def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Жезлы и посохи ювелирные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/products/yuvelirnaya-produktsiya/zhezly-i-posohi-yuvelirnye?group_id=296&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=328&shopFilter%5Bprice_from%5D=240&shopFilter%5Bprice_to%5D=809500&shopFilter%5Bpc%5D=n&page={}'.format(i))


def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'звездицы '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%B7%D0%B2%D0%B5%D0%B7%D0%B4%D0%B8%D1%86%D1%8B')


def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = ' Иконы и складни '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,6):
    pos_parse_4('https://sofrino.ru/products/yuvelirnaya-produktsiya/ikony-i-skladni-yuvelirnye?group_id=296&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=342&shopFilter%5Bprice_from%5D=240&shopFilter%5Bprice_to%5D=809500&shopFilter%5Bpc%5D=n&page={}'.format(i))


def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = ' Иконы нательные ювелирные  '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/products/yuvelirnaya-produktsiya/ikony-natelnye-yuvelirnye?group_id=342&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=523&shopFilter%5Bprice_from%5D=240&shopFilter%5Bprice_to%5D=809500&shopFilter%5Bpc%5D=n&page={}'.format(i))



def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Кадила ювелирные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/products/yuvelirnaya-produktsiya/kadila-yuvelirnye?group_id=523&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=348&shopFilter%5Bprice_from%5D=1350&shopFilter%5Bprice_to%5D=269750&shopFilter%5Bpc%5D=n&page={}'.format(i))







def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Ковши ювелирные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/yuvelirnaya-produktsiya?phrase=%D0%9A%D0%BE%D0%B2%D1%88%D0%B8')



def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Копие ювелирные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/yuvelirnaya-produktsiya?phrase=%D0%9A%D0%BE%D0%BF%D0%B8%D0%B5')






def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Кресты на клобук '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/yuvelirnaya-produktsiya?phrase=%D0%9A%D1%80%D0%B5%D1%81%D1%82%D1%8B+%D0%BD%D0%B0+%D0%BA%D0%BB%D0%BE%D0%B1%D1%83%D0%BA')








def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Кресты на митру'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2/yuvelirnaya-produktsiya?phrase=%D0%9A%D1%80%D0%B5%D1%81%D1%82%D1%8B+%D0%BD%D0%B0+%D0%BC%D0%B8%D1%82%D1%80%D1%83&page={}'.format(i))













def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Кресты наперсные ювелирные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,16):
    pos_parse_4('https://sofrino.ru/products/yuvelirnaya-produktsiya/kresty-napersnye-yuvelirnye?group_id=296&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=334&shopFilter%5Bprice_from%5D=240&shopFilter%5Bprice_to%5D=809500&shopFilter%5Bpc%5D=n&page={}'.format(i))








def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Кресты напрестольные ювелирные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/products/yuvelirnaya-produktsiya/kresty-naprestolnye-yuvelirnye?group_id=334&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=335&shopFilter%5Bprice_from%5D=9600&shopFilter%5Bprice_to%5D=500700&shopFilter%5Bpc%5D=n&page={}'.format(i))







def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Кресты нательные ювелирные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,5):
    pos_parse_4('https://sofrino.ru/products/yuvelirnaya-produktsiya/kresty-natelnye-yuvelirnye?group_id=335&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=355&shopFilter%5Bprice_from%5D=240&shopFilter%5Bprice_to%5D=809500&shopFilter%5Bpc%5D=n&page={}'.format(i))








def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Лампады ювелирные '

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,4):
    pos_parse_4('https://sofrino.ru/products/yuvelirnaya-produktsiya/lampady-yuvelirnye?group_id=355&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=336&shopFilter%5Bprice_from%5D=240&shopFilter%5Bprice_to%5D=809500&shopFilter%5Bpc%5D=n&page={}'.format(i))














def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Лампада подвесная'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/products/yuvelirnaya-produktsiya/lampady-yuvelirnye?group_id=336&tag_id=0&is_search=&phrase=&shopFilter%5Bgroups%5D%5B0%5D=336&shopFilter%5Bprops%5D%5B168%5D%5B0%5D=2858&shopFilter%5Bprice_from%5D=28000&shopFilter%5Bprice_to%5D=318900&shopFilter%5Bpc%5D=n&page={}'.format(i))













def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Лжицы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2/yuvelirnaya-produktsiya?phrase=%D0%9B%D0%B6%D0%B8%D1%86%D1%8B&page={}'.format(i))















def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Наборы'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/yuvelirnaya-produktsiya?phrase=%D0%BD%D0%B0%D0%B1%D0%BE%D1%80')






















def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Панагии'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,15):
    pos_parse_4('https://sofrino.ru/poisk2/yuvelirnaya-produktsiya?phrase=%D0%9F%D0%B0%D0%BD%D0%B0%D0%B3%D0%B8%D0%B8&page={}'.format(i))













def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Подсвечники'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/yuvelirnaya-produktsiya?phrase=%D0%9F%D0%BE%D0%B4%D1%81%D0%B2%D0%B5%D1%87%D0%BD%D0%B8%D0%BA%D0%B8')








def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Потир-чаши'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,10):
    pos_parse_4('https://sofrino.ru/poisk2/yuvelirnaya-produktsiya?phrase=%D0%9F%D0%BE%D1%82%D0%B8%D1%80-%D1%87%D0%B0%D1%88%D0%B8&page={}'.format(i))





def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Разное'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/products/tags/novinki/yuvelirnaya-produktsiya')







def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Рамки на митру'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?phrase=%D0%A0%D0%B0%D0%BC%D0%BA%D0%B8%20%D0%BD%D0%B0%20%D0%BC%D0%B8%D1%82%D1%80%D1%83')







def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Сосуды'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/yuvelirnaya-produktsiya?phrase=%D0%A1%D0%BE%D1%81%D1%83%D0%B4%D1%8B')






def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Сосуды'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2/yuvelirnaya-produktsiya?phrase=%D0%A1%D0%BE%D1%81%D1%83%D0%B4%D1%8B')









def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Цепи для крестов и панагий'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,2):
    pos_parse_4('https://sofrino.ru/poisk2?group_id=0&tag_id=0&is_search=1&phrase=%D0%A6%D0%B5%D0%BF%D0%B8+%D0%B4%D0%BB%D1%8F+%D0%BA%D1%80%D0%B5%D1%81%D1%82%D0%BE%D0%B2+%D0%B8+%D0%BF%D0%B0%D0%BD%D0%B0%D0%B3%D0%B8%D0%B9&shopFilter%5Bprops%5D%5B156%5D%5B%5D=980&shopFilter%5Bprice_from%5D=&shopFilter%5Bprice_to%5D=&shopFilter%5Bpc%5D=n')






def pos_check_4():
    global g, ids
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    g+=2
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]

    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    top = sheet['C{}'.format(g)]
    top.value = 'Яйца пасхальные'

    wb.save('output.xlsx')
def hunter():
    global ids
    ids = 0
def pos_parse_4(url):
    global iconostasi, g, ids
    print(url)
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    HEADERS = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; rv:78.0) Gecko/20100101 Firefox/78.0' }
    response = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(response.content, 'html.parser')
    items = soup.findAll('div',class_='col-4 product-box')
    iconostasi = []

    for item in items:
        iconostasi.append({
            'title': item.find('a', class_='product__title').get_text(strip = True),
            'articul': item.find('div', class_='product__footer').get_text(strip = True),
            'price': item.find('div', class_='product__price__current').get_text(strip = True)
        })

    for comp in iconostasi:
        ff = []
        ff1 = []
        g += 1
        ids += 1
        ff.append(comp['articul'].split('-'))
        for i in ff[0]:
            ff1.append(i.split(' '))
        print('Articul ', ff1[0][1])
        sheet.cell(row=g, column=1).value = str(ids)
        sheet.cell(row=g,column=2).value= ff1[0][1]
        formatted = 'C{}:F{}'.format(g,g)
        sheet.merge_cells(formatted)
        formatted_2 = 'C{}'.format(g)
        top = sheet[formatted_2]
        top.value = comp['title']
        top = sheet['G{}'.format(g)]
        top.value = 'шт'
        top = sheet['H{}'.format(g)]
        top.value = comp['price']
    formatted = 'C{}:F{}'.format(g,g)
    sheet.merge_cells(formatted)
    wb.save('output.xlsx')
pos_check_4()
hunter()
for i in range(1,3):
    pos_parse_4('https://sofrino.ru/poisk2/yuvelirnaya-produktsiya?phrase=%D0%AF%D0%B9%D1%86%D0%B0+%D0%BF%D0%B0%D1%81%D1%85%D0%B0%D0%BB%D1%8C%D0%BD%D1%8B%D0%B5&page={}'.format(i))



def func():
    wb = load_workbook('output.xlsx')
    print('wb opened')
    sheet = wb['Sheet1']
    m_row = sheet.max_row
    print(m_row)
    for i in range(1, m_row + 1):
        cell_obj = sheet.cell(row=i, column=2)
        obg = []
        try:
            #print(cell_obj.value)
            a = cell_obj.value
            if a != 'Артикул':
                fd = a[0:3]
                gd = a[4:7]
                gg = (a[0:3] ,a[3:7])
                fg = (gg[0], gg[1])
                print(fg)
                sheet.cell(row=i,column=2).value = '{} {}'.format(gg[0],gg[1])
                print('added')
        except:
            continue
    wb.save('output.xlsx')
func()

def func2():
    from openpyxl import Workbook, load_workbook
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    m_row = sheet.max_row
    print(m_row)
    for i in range(1,m_row+1):
        print( sheet.cell(row=i,column=8).value)
        a = sheet.cell(row=i,column=8).value
        try:
            sheet.cell(row=i,column=15).value = a.replace('₽','')
        except:
            continue
    wb.save('output.xlsx')
func2()

def func3():
    wb = load_workbook('output.xlsx')
    sheet = wb['Sheet1']
    m_row = sheet.max_row
    print(m_row)
    for i in range(9,m_row+1):
        k = sheet.cell(row=i,column=2).value
        try:
            if len(k) != 8 :
                print('DEBUG: ',k,'\n',i)
                sheet.cell(row=i,column=2).value = None
                sheet.cell(row=i,column=1).value = None
                sheet.cell(row=i,column=3).value = None
                sheet.cell(row=i,column=4).value = None
                sheet.cell(row=i,column=5).value = None
                sheet.cell(row=i,column=15).value = None
                #time.sleep(2)
        except:
            continue
    wb.save('output.xlsx')
func3()